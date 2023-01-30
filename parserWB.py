import time
import os.path
import requests
import xlsxwriter
from bs4 import BeautifulSoup
from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.common.by import By

headers = {'accept': '*/*',
           'accept-language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
           'content-type': 'application/json',
           'origin': 'https://www.wildberries.ru',
           'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.5005.62 Safari/537.36',
           }


def main(product_id):
    try:
        req_card = f'https://card.wb.ru/cards/detail?spp=0&regions=80,64,83,4,38,33,70,68,69,86,75,30,40,48,1,66,31,22,71&pricemarginCoeff=1.0&reg=0&appType=1&emp=0&locale=ru&lang=ru&curr=rub&couponsGeo=12,3,18,15,21&dest=-1029256,-102269,-2162196,-1257786&nm={product_id}'
        json_card = requests.get(req_card, headers=headers).json()

        imt_id = json_card['data']['products'][0]['root']

        url_get_feedbacks = f'https://feedbacks1.wb.ru/feedbacks/v1/{imt_id}'

        json_feedbacks = requests.get(url_get_feedbacks, headers=headers).json()

        return json_card, json_feedbacks

    except Exception as ex:
        print(ex)


def pars_data(product_id):
    try:
        json_card, json_feedbacks = main(product_id)

        count_feedbacks = json_feedbacks['feedbackCount']

        if count_feedbacks != 0:
            print('Count feedbacks (json) ' + str(count_feedbacks))

            for i in range(count_feedbacks):
                if json_feedbacks['feedbacks'][i]['productValuation'] == 5:
                    name = json_feedbacks['feedbacks'][i]['wbUserDetails']['name']
                    date = json_feedbacks['feedbacks'][i]['createdDate'].replace('T', ', ').replace('Z', ' ').split('.')[0]
                    feedback_text = json_feedbacks['feedbacks'][i]['text']

                    yield name, date, feedback_text

        else:
            req_card = f'https://card.wb.ru/cards/detail?spp=0&regions=80,64,83,4,38,33,70,68,69,86,75,30,40,48,1,66,31,22,71&pricemarginCoeff=1.0&reg=0&appType=1&emp=0&locale=ru&lang=ru&curr=rub&couponsGeo=12,3,18,15,21&dest=-1029256,-102269,-2162196,-1257786&nm={product_id}'
            json_card = requests.get(req_card, headers=headers).json()
            imt_id = json_card['data']['products'][0]['root']

            driver = webdriver.Chrome('webdriver\chromedriver.exe')
            driver.get(url=f'https://www.wildberries.ru/catalog/{product_id}/feedbacks?imtId={imt_id}&size=218740513')
            time.sleep(2)

            body = driver.find_element(By.CSS_SELECTOR, 'div[class="wrapper"]')
            body.click()

            for r in range(35):
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(0.15)

            with open('wb_site.html', 'w', encoding='utf-8') as file:
                file.write(driver.page_source)
            file.close()

            with open('wb_site.html', encoding='utf-8') as file:
                html = file.read()
                soup = BeautifulSoup(html, 'lxml')

            all_feedbacks = soup.find_all('li', class_='comments__item feedback j-feedback-slide')

            for feedback in all_feedbacks:
                count_feedbacks += 1

                feedback_top_wrap = feedback.find('div', class_='feedback__top-wrap')
                feedback_info = feedback_top_wrap.find('div', class_='feedback__wrap')
                feedback_rating = feedback_info.find('span', class_='feedback__rating stars-line star5')

                if feedback_rating is not None:
                    feedback_info = feedback_top_wrap.find('div', class_='feedback__info')
                    name = feedback_info.find('button', class_='feedback__header').text
                    date = feedback_info.find('span', class_='feedback__date hide-mobile').get('content').replace(
                        'T', ', ').replace('Z', '').split('.')[0]
                    feedback_content = feedback.find('div', class_='feedback__content')
                    feedback_text = feedback_content.find('p', class_='feedback__text').text

                    yield name, date, feedback_text

                else:
                    pass

            file.close()

            print('Count feedbacks (html) ' + str(count_feedbacks))

            os.remove('wb_site.html')

    except Exception as ex:
        print(ex)


def start(product_id, id=None):
    file_name = 'feedbacks/feedbacks.xlsx'

    if os.path.exists(file_name):
        xl_file = load_workbook(file_name)
        page = xl_file['feedbacks']
        for data in pars_data(product_id):
            page.append([product_id, data[0], data[1], data[2], id])

        xl_file.save(file_name)
        xl_file.close()

    else:
        book = xlsxwriter.Workbook(file_name)  # Создаем файл Exel
        page = book.add_worksheet('feedbacks')

        row = 0

        page.set_column('A:A', 20)
        page.set_column('B:B', 15)
        page.set_column('C:C', 20)
        page.set_column('D:D', 35)
        page.set_column('E:E', 15)

        page.write(0, 0, 'Артикул WB')
        page.write(0, 0 + 1, 'Имя')
        page.write(0, 0 + 2, 'Дата')
        page.write(0, 0 + 3, 'Отзыв')
        page.write(0, 0 + 4, 'ID на сайте')
        row += 1

        book.close()

        xl_file = load_workbook(file_name)
        page = xl_file['feedbacks']

        for data in pars_data(product_id):
            page.append([product_id, data[0], data[1], data[2], id])

        xl_file.save(file_name)
        xl_file.close()